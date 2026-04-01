#!/usr/bin/env python3
"""
Trend Micro LMPI - controllo overuse licenze.

Obiettivo:
- interrogare il report Customer Summary di LMPI;
- confrontare licenze provisioned vs used per cliente/prodotto/service plan;
- evidenziare tutti i clienti che usano piu' licenze del disponibile;
- esportare il dettaglio dell'eccesso per ogni cliente.

Lo script e' stato strutturato in modo piu' operativo/enterprise, ispirandosi
all'organizzazione del file di esempio caricato dall'utente:
- configurazione centralizzata via .env / CLI
- logging uniforme
- helper functions dedicate
- main() come orchestratore
- output sia CSV sia Excel
"""

from __future__ import annotations

import argparse
import base64
import csv
import datetime as dt
import hashlib
import hmac
import json
import logging
import os
import sys
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import urlencode, urlparse

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv as _dotenv_load
except Exception:  # pragma: no cover - fallback safe
    _dotenv_load = None

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    stream=sys.stdout,
    force=True,
)

DEFAULT_BASE_URL = "https://cspi.trendmicro.com"
DEFAULT_LANGUAGE = "en-US"
DEFAULT_TIMEOUT = 60
DEFAULT_OUTPUT_DIR = "./output"


class LMPIError(RuntimeError):
    """Errore funzionale/tecnico proveniente dal flusso LMPI."""


@dataclass(frozen=True)
class OveruseRow:
    customer: str
    city: str
    state: str
    owned_by_vendor: str
    created_by_vendor: str
    product_name: str
    service_plan: str
    unit: str
    provisioned: int
    used: int
    excess: int


@dataclass(frozen=True)
class AppConfig:
    env_file: str
    base_url: str
    access_token: Optional[str]
    secret_key: Optional[str]
    partner_id: Optional[str]
    product_id: Optional[str]
    language_code: str
    report_year: str
    report_month: str
    output_dir: str
    input_json: Optional[str]
    raw_json: bool
    min_excess: int
    fail_on_overuse: bool
    timeout: int
    write_excel: bool


class LMPIClient:
    def __init__(self, base_url: str, access_token: str, secret_key: str, *, timeout: int = DEFAULT_TIMEOUT) -> None:
        self.base_url = base_url.rstrip("/")
        self.access_token = access_token
        self.secret_key = secret_key
        self.timeout = timeout
        self.session = requests.Session()

    def _build_request_uri(self, path: str, params: Optional[Dict[str, Any]] = None) -> str:
        parsed = urlparse(path)
        request_path = parsed.path if parsed.scheme and parsed.netloc else path
        if params:
            query = urlencode(params, doseq=True, safe=":")
            return f"{request_path}?{query}"
        return request_path

    @staticmethod
    def _md5_base64(content_bytes: bytes) -> str:
        digest = hashlib.md5(content_bytes).digest()
        return base64.b64encode(digest).decode("ascii")

    def _build_signature(self, posix_time: int, method: str, request_uri: str, body_bytes: bytes) -> str:
        message = f"{posix_time}{method.upper()}{request_uri}"
        if body_bytes:
            message += self._md5_base64(body_bytes)

        digest = hmac.new(
            self.secret_key.encode("utf-8"),
            message.encode("utf-8"),
            hashlib.sha256,
        ).digest()
        return base64.b64encode(digest).decode("ascii")

    def request(
        self,
        method: str,
        path: str,
        *,
        params: Optional[Dict[str, Any]] = None,
        payload: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        request_uri = self._build_request_uri(path, params=params)
        url = f"{self.base_url}{request_uri}"

        body_bytes = b""
        if payload is not None:
            body_bytes = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")

        posix_time = int(dt.datetime.now(dt.timezone.utc).timestamp())
        headers = {
            "x-access-token": self.access_token,
            "x-posix-time": str(posix_time),
            "x-signature": self._build_signature(posix_time, method, request_uri, body_bytes),
            "x-traceid": str(uuid.uuid4()),
            "content-type": "application/json; charset=utf-8",
        }

        logging.info("Chiamata API %s %s", method.upper(), request_uri)
        response = self.session.request(
            method=method.upper(),
            url=url,
            params=params,
            data=body_bytes if body_bytes else None,
            headers=headers,
            timeout=self.timeout,
        )

        if response.status_code >= 400:
            raise LMPIError(
                f"HTTP {response.status_code} su {method.upper()} {request_uri}: "
                f"{response.text.strip()[:1000]}"
            )

        if not response.text.strip():
            logging.warning("Risposta vuota da %s %s", method.upper(), request_uri)
            return {}

        try:
            return response.json()
        except ValueError as exc:
            raise LMPIError(
                f"Risposta JSON non valida da {method.upper()} {request_uri}: {response.text[:1000]}"
            ) from exc

    def customer_summary(
        self,
        *,
        year: str,
        month: str,
        language_code: str,
        partner_id: Optional[str] = None,
        product_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        payload: Dict[str, Any] = {
            "report_cycle_year": year,
            "report_cycle_month": month,
            "language_code": language_code,
        }
        if partner_id:
            payload["partner_id"] = partner_id
        if product_id:
            payload["product_id"] = product_id

        return self.request("POST", "/LMPI/v2/reports/summary", payload=payload)


def manual_load_env_file(env_file: str) -> None:
    path = Path(env_file)
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def preload_env_file(env_file: str) -> None:
    if not env_file:
        return

    path = Path(env_file)
    if not path.exists():
        logging.debug("File .env non trovato: %s", env_file)
        return

    if _dotenv_load is not None:
        _dotenv_load(dotenv_path=path, override=False)
    else:
        manual_load_env_file(env_file)


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    pre_parser = argparse.ArgumentParser(add_help=False)
    pre_parser.add_argument("--env-file", default=os.getenv("LMPI_ENV_FILE", ".env"))
    pre_args, _ = pre_parser.parse_known_args(argv)
    preload_env_file(pre_args.env_file)

    now = dt.datetime.now(dt.timezone.utc)

    parser = argparse.ArgumentParser(
        description="Trend Micro LMPI: trova i clienti che usano piu' licenze di quelle disponibili."
    )
    parser.add_argument("--env-file", default=pre_args.env_file, help="File .env da caricare prima del parsing finale")
    parser.add_argument("--base-url", default=os.getenv("LMPI_BASE_URL", DEFAULT_BASE_URL), help="Base URL LMPI")
    parser.add_argument("--access-token", default=os.getenv("LMPI_ACCESS_TOKEN"), help="LMPI access token")
    parser.add_argument("--secret-key", default=os.getenv("LMPI_SECRET_KEY"), help="LMPI secret key")
    parser.add_argument("--partner-id", default=os.getenv("LMPI_PARTNER_ID"), help="Filtro opzionale partner_id")
    parser.add_argument("--product-id", default=os.getenv("LMPI_PRODUCT_ID"), help="Filtro opzionale product_id")
    parser.add_argument(
        "--language-code",
        default=os.getenv("LMPI_LANGUAGE_CODE", DEFAULT_LANGUAGE),
        help="Language code richiesto dall'API",
    )
    parser.add_argument(
        "--year",
        default=os.getenv("LMPI_REPORT_YEAR", f"{now.year:04d}"),
        help="Anno del report cycle",
    )
    parser.add_argument(
        "--month",
        default=os.getenv("LMPI_REPORT_MONTH", f"{now.month:02d}"),
        help="Mese del report cycle",
    )
    parser.add_argument(
        "--output-dir",
        default=os.getenv("LMPI_OUTPUT_DIR", DEFAULT_OUTPUT_DIR),
        help="Directory di output",
    )
    parser.add_argument("--input-json", help="Usa una risposta JSON gia' salvata invece di chiamare l'API")
    parser.add_argument("--raw-json", action="store_true", help="Salva anche il JSON grezzo della risposta")
    parser.add_argument(
        "--min-excess",
        type=int,
        default=int(os.getenv("LMPI_MIN_EXCESS", "1")),
        help="Soglia minima di eccesso per includere una riga nel report",
    )
    parser.add_argument(
        "--fail-on-overuse",
        action="store_true",
        help="Esce con codice 2 se trova almeno un overuse",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=int(os.getenv("LMPI_TIMEOUT", str(DEFAULT_TIMEOUT))),
        help="Timeout HTTP in secondi",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Disabilita la generazione del file Excel",
    )
    return parser.parse_args(argv)


def build_config(args: argparse.Namespace) -> AppConfig:
    return AppConfig(
        env_file=args.env_file,
        base_url=str(args.base_url).strip(),
        access_token=_none_if_empty(args.access_token),
        secret_key=_none_if_empty(args.secret_key),
        partner_id=_none_if_empty(args.partner_id),
        product_id=_none_if_empty(args.product_id),
        language_code=str(args.language_code).strip() or DEFAULT_LANGUAGE,
        report_year=str(args.year).strip(),
        report_month=str(args.month).strip().zfill(2),
        output_dir=str(args.output_dir).strip() or DEFAULT_OUTPUT_DIR,
        input_json=_none_if_empty(args.input_json),
        raw_json=bool(args.raw_json),
        min_excess=int(args.min_excess),
        fail_on_overuse=bool(args.fail_on_overuse),
        timeout=int(args.timeout),
        write_excel=not bool(args.no_excel),
    )


def _none_if_empty(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def require(value: Optional[str], label: str) -> str:
    if value:
        return value
    raise SystemExit(f"Parametro obbligatorio mancante: {label}")


def to_int(value: Any, *, default: int = 0) -> int:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return int(value)
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(str(value).replace(",", ".")))
        except (TypeError, ValueError):
            return default


def safe_filename(value: str) -> str:
    cleaned = "".join(char if char.isalnum() or char in {"-", "_", "."} else "_" for char in value)
    cleaned = cleaned.strip("_")
    return cleaned or "report"


def ensure_output_dir(path_str: str) -> Path:
    path = Path(path_str)
    path.mkdir(parents=True, exist_ok=True)
    return path


def response_summary_rows(response: Dict[str, Any]) -> List[Dict[str, Any]]:
    summary_rows = response.get("summary")
    if isinstance(summary_rows, list):
        return summary_rows

    raise LMPIError(
        "Struttura risposta non attesa: chiave 'summary' assente oppure non lista. "
        f"Chiavi disponibili: {', '.join(sorted(response.keys()))}"
    )


def normalize_overuse_rows(summary_rows: Iterable[Dict[str, Any]], *, min_excess: int) -> List[OveruseRow]:
    overuse_rows: List[OveruseRow] = []

    for row in summary_rows:
        used_raw = row.get("used")
        if used_raw is None:
            continue

        provisioned = to_int(row.get("provisioned"))
        used = to_int(used_raw)
        excess = used - provisioned
        if excess < min_excess:
            continue

        overuse_rows.append(
            OveruseRow(
                customer=str(row.get("customer", "")).strip() or "<unknown>",
                city=str(row.get("city", "")).strip(),
                state=str(row.get("state", "")).strip(),
                owned_by_vendor=str(row.get("owned_by_vendor", "")).strip(),
                created_by_vendor=str(row.get("created_by_vendor", "")).strip(),
                product_name=str(row.get("product_name", "")).strip(),
                service_plan=str(row.get("service_plan", "")).strip(),
                unit=str(row.get("unit", "")).strip() or "Seats",
                provisioned=provisioned,
                used=used,
                excess=excess,
            )
        )

    overuse_rows.sort(key=lambda item: (-item.excess, -item.used, item.customer.lower(), item.product_name.lower()))
    return overuse_rows


def aggregate_by_customer(rows: Iterable[OveruseRow]) -> List[Dict[str, Any]]:
    aggregated: Dict[str, Dict[str, Any]] = {}

    for row in rows:
        item = aggregated.setdefault(
            row.customer,
            {
                "customer": row.customer,
                "city": row.city,
                "state": row.state,
                "owned_by_vendor": row.owned_by_vendor,
                "created_by_vendor": row.created_by_vendor,
                "total_excess": 0,
                "overused_products": 0,
                "details": [],
            },
        )
        item["total_excess"] += row.excess
        item["overused_products"] += 1
        item["details"].append(f"{row.product_name} / {row.service_plan}: +{row.excess} {row.unit}")

    result = list(aggregated.values())
    result.sort(key=lambda item: (-item["total_excess"], item["customer"].lower()))
    return result


def overuse_rows_to_dicts(rows: Iterable[OveruseRow]) -> List[Dict[str, Any]]:
    return [
        {
            "customer": row.customer,
            "city": row.city,
            "state": row.state,
            "owned_by_vendor": row.owned_by_vendor,
            "created_by_vendor": row.created_by_vendor,
            "product_name": row.product_name,
            "service_plan": row.service_plan,
            "unit": row.unit,
            "provisioned": row.provisioned,
            "used": row.used,
            "excess": row.excess,
        }
        for row in rows
    ]


def aggregated_rows_to_dicts(rows: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "customer": row["customer"],
            "city": row["city"],
            "state": row["state"],
            "owned_by_vendor": row["owned_by_vendor"],
            "created_by_vendor": row["created_by_vendor"],
            "total_excess": row["total_excess"],
            "overused_products": row["overused_products"],
            "details": " | ".join(row["details"]),
        }
        for row in rows
    ]


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    if rows:
        fieldnames = list(rows[0].keys())
    else:
        fieldnames = ["message"]
        rows = [{"message": "no_data"}]

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def autosize_worksheet(worksheet) -> None:
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 80)


def write_excel_report(path: Path, detail_rows: List[Dict[str, Any]], customer_rows: List[Dict[str, Any]]) -> None:
    workbook = Workbook()

    details_sheet = workbook.active
    details_sheet.title = "Overuse Details"
    _write_sheet_rows(details_sheet, detail_rows)
    autosize_worksheet(details_sheet)

    summary_sheet = workbook.create_sheet(title="By Customer")
    _write_sheet_rows(summary_sheet, customer_rows)
    autosize_worksheet(summary_sheet)

    workbook.save(path)


def _write_sheet_rows(worksheet, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        worksheet.append(["message"])
        worksheet.append(["no_data"])
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])


def print_terminal_report(detail_rows: List[OveruseRow], customer_rows: List[Dict[str, Any]]) -> None:
    print()
    print("=== Trend Micro LMPI - License Overuse Report ===")
    print(f"Clienti con overuse: {len(customer_rows)}")
    print(f"Righe cliente/prodotto in overuse: {len(detail_rows)}")
    print()

    if not detail_rows:
        print("Nessun overuse trovato per il periodo selezionato.")
        return

    print("Top clienti per eccesso totale:")
    for item in customer_rows[:20]:
        print(
            f"- {item['customer']}: +{item['total_excess']} "
            f"su {item['overused_products']} combinazioni prodotto/piano"
        )

    print()
    print("Dettaglio prime righe:")
    for row in detail_rows[:50]:
        print(
            f"- {row.customer} | {row.product_name} | {row.service_plan} | "
            f"provisioned={row.provisioned} | used={row.used} | excess=+{row.excess} {row.unit}"
        )
    if len(detail_rows) > 50:
        print(f"... console troncata: altre {len(detail_rows) - 50} righe disponibili nei file di output.")


def load_response(config: AppConfig) -> Dict[str, Any]:
    if config.input_json:
        logging.info("Caricamento risposta da file JSON: %s", config.input_json)
        return json.loads(Path(config.input_json).read_text(encoding="utf-8"))

    access_token = require(config.access_token, "LMPI_ACCESS_TOKEN / --access-token")
    secret_key = require(config.secret_key, "LMPI_SECRET_KEY / --secret-key")

    client = LMPIClient(
        base_url=config.base_url,
        access_token=access_token,
        secret_key=secret_key,
        timeout=config.timeout,
    )
    return client.customer_summary(
        year=config.report_year,
        month=config.report_month,
        language_code=config.language_code,
        partner_id=config.partner_id,
        product_id=config.product_id,
    )


def save_outputs(
    config: AppConfig,
    response: Dict[str, Any],
    detail_rows: List[OveruseRow],
    customer_rows: List[Dict[str, Any]],
) -> Dict[str, Path]:
    output_dir = ensure_output_dir(config.output_dir)
    suffix = f"{config.report_year}{config.report_month}"

    detail_dicts = overuse_rows_to_dicts(detail_rows)
    customer_dicts = aggregated_rows_to_dicts(customer_rows)

    detail_csv_path = output_dir / safe_filename(f"overuse_details_{suffix}.csv")
    customer_csv_path = output_dir / safe_filename(f"overuse_by_customer_{suffix}.csv")

    write_csv(detail_csv_path, detail_dicts)
    write_csv(customer_csv_path, customer_dicts)

    saved_paths: Dict[str, Path] = {
        "detail_csv": detail_csv_path,
        "customer_csv": customer_csv_path,
    }

    if config.write_excel:
        excel_path = output_dir / safe_filename(f"overuse_report_{suffix}.xlsx")
        write_excel_report(excel_path, detail_dicts, customer_dicts)
        saved_paths["excel"] = excel_path

    if config.raw_json:
        raw_json_path = output_dir / safe_filename(f"raw_customer_summary_{suffix}.json")
        raw_json_path.write_text(json.dumps(response, ensure_ascii=False, indent=2), encoding="utf-8")
        saved_paths["raw_json"] = raw_json_path

    return saved_paths


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    config = build_config(args)

    logging.info(
        "Avvio controllo overuse: year=%s month=%s partner_id=%s product_id=%s input_json=%s",
        config.report_year,
        config.report_month,
        config.partner_id or "<all>",
        config.product_id or "<all>",
        config.input_json or "<api>",
    )

    response = load_response(config)
    summary_rows = response_summary_rows(response)
    detail_rows = normalize_overuse_rows(summary_rows, min_excess=config.min_excess)
    customer_rows = aggregate_by_customer(detail_rows)
    saved_paths = save_outputs(config, response, detail_rows, customer_rows)

    print_terminal_report(detail_rows, customer_rows)
    print()
    print(f"CSV dettaglio:  {saved_paths['detail_csv']}")
    print(f"CSV clienti:    {saved_paths['customer_csv']}")
    if "excel" in saved_paths:
        print(f"Excel report:   {saved_paths['excel']}")
    if "raw_json" in saved_paths:
        print(f"JSON grezzo:    {saved_paths['raw_json']}")

    if config.fail_on_overuse and detail_rows:
        logging.warning("Overuse rilevato: uscita con codice 2")
        return 2

    logging.info("Elaborazione completata con successo")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        raise SystemExit("Interrotto dall'utente")
    except LMPIError as exc:
        raise SystemExit(f"Errore LMPI: {exc}")
